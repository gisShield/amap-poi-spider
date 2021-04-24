import logging
import sys
import traceback
import openpyxl
import requests
import geocoding
import os

LOG_FORMAT = "%(asctime)s %(name)s %(levelname)s %(pathname)s %(message)s "  # 配置输出日志格式
DATE_FORMAT = '%Y-%m-%d  %H:%M:%S %a '  # 配置输出时间的格式，注意月份和天数不要搞乱了
logging.basicConfig(level=logging.INFO,
                    format=LOG_FORMAT,
                    datefmt=DATE_FORMAT,
                    filename=r"./logs/poiSpider.log"  # 有了filename参数就不会直接输出显示到控制台，而是直接写入文件
                    )


class PoiSpider:
    def __init__(self, key, polygon_points, type_code, keywords):
        # 接口地址
        self.spider_url = 'https://restapi.amap.com/v3/place/polygon'
        # 高德api key
        self.api_key_list = key

        # POI 大类列表 参考：https://lbs.amap.com/api/webservice/download 的 POI分类编码
        self.type_code_list = str.split(type_code, '|')
        self.polygon_points = polygon_points
        self.keywords = keywords
        # self.city_code = area_code
        self.fileName = 'POI爬取.xlsx'
        self.workbook = None
        self.res_list = []

    def get_poi(self):
        """
        获取POI信息
        """
        # 获取大类
        for code in self.type_code_list:
            self.get_json(poi_type=code, page=1)

    def get_json(self, poi_type, page):
        """
        发送请求获取结果集
        :param poi_type: 字典类型
        :param page: 当前页数
        :return:
        """
        try:
            url_params = {
                'polygon': self.polygon_points,
                # 'city': self.city_code,
                'key': self.api_key_list,
                'offset': 20,
                'page': page,
                'types': poi_type
            }
            if self.keywords:
                url_params.keywords = self.keywords
            print('请求 %s %d 页数据开始' % (poi_type, page))

            requests.adapters.DEFAULT_RETRIES = 5
            s = requests.session()
            s.keep_alive = False
            res = requests.get(self.spider_url, headers={'Connection': 'close'}, params=url_params)
            if res.status_code == 200:
                res_dict = res.json()
                print(res_dict['count'])
                # 获取POI的信息
                pois = res_dict['pois']
                for poi in pois:
                    # 避免值出现空[]等错误，进行判断后再加入列表
                    poi_list = [
                        poi['id'] if "id" in poi and poi['id'] else '',
                        poi['name'] if "name" in poi and poi['name'] else '',
                        poi['type'] if "type" in poi and poi['type'] else '',
                        self.change_to_wgs84(poi['location']) if "location" in poi and poi['location'] else '',
                        poi['tel'] if "tel" in poi and poi['tel'] else '',
                        poi['pname'] if "pname" in poi and poi['pname'] else '',
                        poi['cityname'] if "cityname" in poi and poi['cityname'] else '',
                        poi['adname'] if "adname" in poi and poi['adname'] else '',
                        poi['address'] if "address" in poi and poi['address'] else ''
                    ]
                    self.res_list.append(poi_list)

                if page * url_params.get('offset') < int(res_dict['count']):
                    self.get_json(poi_type=poi_type, page=page + 1)
                else:
                    self.save_poi(poi_type)
                return 0
            else:
                self.save_poi(poi_type)
                print('接口请求异常！错误信息为：%s' % res.json())
                print("当前正在请求的接口参数为：{page: %s, types: %s}  !" % (url_params.get('page'), url_params.get('types')))
                print("输入任意字符程序关闭...")
                input()
                return -1
        except Exception as e:
            logging.error("程序异常====> %s", e.args)
            logging.error("异常信息====> %s", traceback.format_exc())
            print('程序出现异常！具体可以查看日志')
        finally:
            self.save_poi(poi_type)

    def change_to_wgs84(self, coord):
        """
        执行坐标转换
        :param coord:坐标
        :return:
        """
        if coord:
            lon = float(coord.split(',')[0])
            lat = float(coord.split(',')[1])
            lonlat = geocoding.gcj02_to_wgs84(lon, lat)
            return str(lonlat[0]) + ',' + str(lonlat[1])
        else:
            return ''

    def save_poi(self, type):
        """
        保存POI信息
        :param type: POI类型
        :return:
        """
        try:
            # 创建文件
            if os.path.isfile(self.fileName):
                self.workbook = openpyxl.load_workbook(self.fileName)
                if type not in self.workbook.sheetnames:
                    self.workbook.create_sheet(type)
                    table = self.workbook[type]
                    # 设置表头
                    table.append(['id', 'name', 'type', 'location', 'tel', 'pname', 'cityname', 'adname', 'address'])
                else:
                    table = self.workbook[type]
            else:
                self.workbook = openpyxl.Workbook()
                table = self.workbook.active
                table.title = type

                # 设置表头
                table.append(['id', 'name', 'type', 'location', 'tel', 'pname', 'cityname', 'adname', 'address'])
                self.workbook.save(self.fileName)

            # 循环插入数据
            for item in self.res_list:
                if item:
                    table.append(item)
            # 保存文件
            self.workbook.save(self.fileName)
            self.workbook.close()  # 对程序中只读的workbook
        except Exception as e:
            logging.error("程序异常====> %s", e.args)
            logging.error("异常信息====> %s", traceback.format_exc())
            print('程序出现异常！具体可以查看日志')
        finally:
            # 清空
            self.res_list = []
        return ''


if __name__ == "__main__":
    try:
        print("1.请输入高德API key")
        key = input()
        if key is None or key == '':
            print("输入参数错误，输入任意字符退出...")
            input()
            exit(0)
        print("2.请输入爬取范围坐标串，经度在前，纬度在后，坐标对用'|'分割。经纬度小数点后不得超过6位,首尾坐标对需相同")
        polygon_points = input()
        if polygon_points is None or polygon_points == '':
            print("参数错误！爬取的范围未输入！任意字符退出...")
            temp = input()
            sys.exit(0)
        print("3.请输入分类代码，多个类型用'|'分割")
        type_code = input()
        if type_code is None or type_code == '':
            print("参数错误！分类代码未输入！任意字符退出...")
            temp = input()
            sys.exit(0)
        print("4.请输入关键字，多个关键字用“|”分割;如果不需要，回车进入下一步")
        keywords = input()
        poi_spider = PoiSpider(key, polygon_points, type_code, keywords)
        poi_spider.get_poi()
        print("爬取完毕！输入任意字符程序关闭...")
        input()
        sys.exit(0)

    except Exception as e:
        logging.error("程序异常====> %s", e.args)
        logging.error("异常信息====> %s", traceback.format_exc())
        print("程序出现异常，输入任意字符退出...")
        input()
        sys.exit(0)

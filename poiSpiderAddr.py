import traceback
import openpyxl
import requests
from openpyxl.utils import column_index_from_string
import logging
import geocoding
import os


class PoiSpiderAddr:
    def __init__(self, key, area_code, input_type, **kw):
        # 接口地址
        self.spider_url = 'https://restapi.amap.com/v3/geocode/geo'
        # 高德api key
        self.api_key_list = key

        self.input_type = input_type

        self.city_code = area_code
        # 属性设置
        for k, w in kw.items():
            setattr(self, k, w)

        if input_type == '2':
            if self.file_path:
                # 文件路径提取 文件名提取
                self.path, self.fileName = os.path.split(self.file_path)

        self.workbook = None

    def get_poi_addr(self):
        """
        读取输入的地址字段
        """
        try:
            self.get_json(address=self.addr, row=-1, sheet=None)
        except Exception as e:
            logging.error("程序异常====> %s", e.args)
            logging.error("异常信息====> %s", traceback.format_exc())
            print('程序出现异常！具体可以查看日志')

    def get_poi_file(self):
        """
        读取Excel表格中字段
        """
        new_path = os.path.join(self.path, 'new_' + self.fileName)
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            sheet = self.workbook[self.sheet_name]  # 获取工作表
            column = sheet[self.addr_col]  # 获取地址这一项
            max_row = sheet.max_row
            for cell in column:
                if self.start_row <= cell.row <= max_row and cell.value:
                    self.get_json(address=cell.value, row=cell.row, sheet=sheet)

            logging.info("坐标爬取成功，新文件路径存放: %s", new_path)
        except Exception as e:
            logging.error("程序异常====> %s", e.args)
            logging.error("异常信息====> %s", traceback.format_exc())
            logging.info("坐标爬取中断，已抓取坐标新文件路径存放: %s", new_path)
            print('程序出现异常！具体可以查看日志')
        finally:
            self.workbook.save(new_path)

    def get_json(self, address, row, sheet):
        """
        发送请求获取结果集
        :param address: 地址
        :param row: 行号
        :param sheet: 工作表
        :return:
        """
        url_params = {
            'city': self.city_code,
            'key': self.api_key_list,
            'address': address
        }
        logging.info('请求 %s 数据开始', address)
        try:
            requests.adapters.DEFAULT_RETRIES = 5
            s = requests.session()
            s.keep_alive = False
            res = requests.get(self.spider_url, headers={'Connection': 'close'}, params=url_params)
            if res.status_code == 200:
                res_dict = res.json()
                logging.info(res_dict['count'])
                if int(res_dict['count']) > 0:
                    # 获取坐标的信息
                    geocodes = res_dict['geocodes']
                    location = self.change_to_wgs84(geocodes[0]['location'])
                    if row == -1 and sheet is None:
                        logging.info('该地址经纬度为：%s', location)
                    else:
                        if self.coord_col.count(',') == 1:
                            sheet.cell(row=row, column=column_index_from_string(self.coord_col.split(',')[0]),
                                       value=location[0])
                            sheet.cell(row=row, column=column_index_from_string(self.coord_col.split(',')[1]),
                                       value=location[1])
                        else:
                            sheet.cell(row=row, column=column_index_from_string(self.coord_col),
                                       value=",".join(location))
            else:
                logging.error('接口请求异常！错误信息为：%s', res.json())
                print('程序出现异常！具体可以查看日志')
                return -1
        except Exception as e:
            raise e

    def change_to_wgs84(self, coord):
        """
        执行坐标转换
        :param coord: 坐标串
        :return:
        """
        if coord:
            lon = float(coord.split(',')[0])
            lat = float(coord.split(',')[1])
            lonlat = geocoding.gcj02_to_wgs84(lon, lat)
            return [str(lonlat[0]), str(lonlat[1])]
        else:
            return ''

# if __name__ == "__main__":
#     try:
#         print("请输入数据所在区划代码 如：330100")
#         area_code = input()
#         print("请选择输入类型(1:参数输入；2：文件输入)")
#         input_type = input()
#         if input_type == '1':
#             print("请输入标准化地址如：浙江省杭州市滨江区星耀城一期")
#             print("注：地址越标准，匹配到的POI信息越精准")
#             addr = input()
#             poi_spider = PoiSpiderAddr(area_code, input_type, addr=addr)
#             poi_spider.get_poi_addr()
#
#         elif input_type == '2':
#             print("请输入文件地址 如：c:\\1.xlsx")
#             file_path = input()
#             print("请输入工作表名称")
#             sheet_name = input()
#             print("请输入数据起始行行号")
#             start_row = int(input())
#             print("请输入地址列列号")
#             addr_col = input()
#             print("请输入经纬度填写列号")
#             coord_col = input()
#             poi_spider = PoiSpiderAddr(area_code, input_type, file_path=file_path, sheet_name=sheet_name,
#                                        start_row=start_row,
#                                        addr_col=addr_col, coord_col=coord_col)
#             poi_spider.get_poi_file()
#         else:
#             logging.error("输入参数错误，程序退出...")
#             exit(0)
#     except Exception as e:
#         logging.error("程序异常====> %s", e.args)
#         logging.error("异常信息====> %s", traceback.format_exc())
